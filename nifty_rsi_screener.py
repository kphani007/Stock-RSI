"""
NIFTY Daily RSI Screener Agent
------------------------------
Runs a daily scan of Indian (NSE) stocks shortly after market open,
finds stocks with 14-day RSI >= 65, and writes a formatted Excel report:

    Date | Stock Name | Current Price | 52 Week High | RSI |
    Recommendation to buy or sell | 1 Year Target

Data source: Yahoo Finance (via the `yfinance` library). NSE tickers on
Yahoo carry the ".NS" suffix (e.g. RELIANCE.NS).

Setup (one time):
    pip install yfinance openpyxl pandas

Run manually:
    python nifty_rsi_screener.py

Schedule daily (see README section at the bottom of this file).

Notes:
- "Recommendation" is the analyst consensus published on Yahoo Finance
  (strong_buy / buy / hold / sell). It is NOT financial advice.
- "1 Year Target" is the analyst mean target price from Yahoo Finance.
- RSI >= 65 generally indicates strong upward momentum approaching
  "overbought" territory (classically 70+). Momentum traders treat this
  as strength; mean-reversion traders treat it as a caution signal.
"""

from __future__ import annotations

import datetime as dt
import os
import sys
import time

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------- configuration -----------------------------

RSI_PERIOD = 14
RSI_THRESHOLD = 65.0
OUTPUT_DIR = os.path.expanduser("~/rsi_reports")  # change if you like
TICKER_FILE = "tickers.csv"  # optional: one Yahoo ticker per line, overrides the list below

# NIFTY 50 universe (Yahoo Finance NSE symbols). Extend freely, or drop a
# tickers.csv next to this script to screen any custom universe
# (e.g. NIFTY 200 / 500 lists downloadable from niftyindices.com).
NIFTY50 = [
    "ADANIENT.NS", "ADANIPORTS.NS", "APOLLOHOSP.NS", "ASIANPAINT.NS",
    "AXISBANK.NS", "BAJAJ-AUTO.NS", "BAJFINANCE.NS", "BAJAJFINSV.NS",
    "BEL.NS", "BHARTIARTL.NS", "CIPLA.NS", "COALINDIA.NS", "DRREDDY.NS",
    "EICHERMOT.NS", "GRASIM.NS", "HCLTECH.NS", "HDFCBANK.NS",
    "HDFCLIFE.NS", "HEROMOTOCO.NS", "HINDALCO.NS", "HINDUNILVR.NS",
    "ICICIBANK.NS", "INDUSINDBK.NS", "INFY.NS", "ITC.NS", "JSWSTEEL.NS",
    "KOTAKBANK.NS", "LT.NS", "M&M.NS", "MARUTI.NS", "NESTLEIND.NS",
    "NTPC.NS", "ONGC.NS", "POWERGRID.NS", "RELIANCE.NS", "SBILIFE.NS",
    "SBIN.NS", "SHRIRAMFIN.NS", "SUNPHARMA.NS", "TATACONSUM.NS",
    "TATAMOTORS.NS", "TATASTEEL.NS", "TCS.NS", "TECHM.NS", "TITAN.NS",
    "TRENT.NS", "ULTRACEMCO.NS", "WIPRO.NS",
]

# ------------------------------- indicators -------------------------------


def compute_rsi(close: pd.Series, period: int = RSI_PERIOD) -> float | None:
    """14-day RSI using Wilder's smoothing (the standard charting formula)."""
    close = close.dropna()
    if len(close) < period + 1:
        return None
    delta = close.diff()
    gain = delta.clip(lower=0.0)
    loss = -delta.clip(upper=0.0)
    avg_gain = gain.ewm(alpha=1.0 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1.0 / period, min_periods=period, adjust=False).mean()
    last_gain = avg_gain.iloc[-1]
    last_loss = avg_loss.iloc[-1]
    if pd.isna(last_gain) or pd.isna(last_loss):
        return None
    if last_loss == 0:
        return 100.0
    rs = last_gain / last_loss
    return float(100.0 - (100.0 / (1.0 + rs)))


# --------------------------------- agent ----------------------------------


def load_universe() -> list[str]:
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), TICKER_FILE)
    if os.path.exists(path):
        with open(path) as f:
            tickers = [line.strip() for line in f if line.strip() and not line.startswith("#")]
        print(f"Loaded {len(tickers)} tickers from {TICKER_FILE}")
        return tickers
    return NIFTY50


def scan(tickers: list[str]) -> list[dict]:
    """Download 1y of daily candles in one batch, then screen for RSI >= threshold."""
    print(f"Downloading price history for {len(tickers)} stocks...")
    data = yf.download(
        tickers, period="1y", interval="1d",
        group_by="ticker", auto_adjust=False, threads=True, progress=False,
    )

    hits: list[dict] = []
    for symbol in tickers:
        try:
            df = data[symbol] if len(tickers) > 1 else data
            close = df["Close"].dropna()
            if close.empty:
                continue
            rsi = compute_rsi(close)
            if rsi is None or rsi < RSI_THRESHOLD:
                continue

            current_price = float(close.iloc[-1])
            week52_high = float(df["High"].dropna().max())

            # Analyst consensus + 1y mean target (best effort; may be missing)
            name, reco, target = symbol.replace(".NS", ""), "n/a", None
            try:
                info = yf.Ticker(symbol).info
                name = info.get("shortName") or name
                reco = (info.get("recommendationKey") or "n/a").replace("_", " ").title()
                target = info.get("targetMeanPrice")
                time.sleep(0.3)  # be polite to the API
            except Exception as exc:  # keep the row even if metadata fails
                print(f"  [warn] metadata failed for {symbol}: {exc}")

            hits.append({
                "date": dt.date.today().isoformat(),
                "stock": name,
                "symbol": symbol.replace(".NS", ""),
                "price": round(current_price, 2),
                "high52": round(week52_high, 2),
                "rsi": round(rsi, 1),
                "reco": reco,
                "target": round(float(target), 2) if target else "n/a",
            })
            print(f"  HIT  {symbol:<15} RSI={rsi:5.1f}  price={current_price:,.2f}")
        except Exception as exc:
            print(f"  [warn] skipped {symbol}: {exc}")

    hits.sort(key=lambda r: r["rsi"], reverse=True)
    return hits


# --------------------------------- excel ----------------------------------

HEADERS = [
    "Date", "Stock Name", "Current Price (₹)", "52 Week High (₹)", "RSI",
    "Recommendation to buy or sell", "1 Year Target (₹)",
]


def write_excel(rows: list[dict]) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, f"RSI_Screener_{dt.date.today().isoformat()}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "RSI Screener"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=11)
    hot_fill = PatternFill("solid", fgColor="FCE4D6")  # RSI >= 70 (overbought)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r["date"], f'{r["stock"]} ({r["symbol"]})', r["price"],
            r["high52"], r["rsi"], r["reco"], r["target"],
        ])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.font = body_font
        for col in (3, 4, 7):
            c = ws.cell(row=row_idx, column=col)
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00"
        ws.cell(row=row_idx, column=5).number_format = "0.0"
        if r["rsi"] >= 70:
            for cell in ws[row_idx]:
                cell.fill = hot_fill

    widths = [12, 34, 16, 16, 8, 26, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(ws.max_row, 2)}"

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1,
            value=f"Screen: 14-day RSI >= {RSI_THRESHOLD:.0f}. Shaded rows: RSI >= 70 (overbought). "
                  "Recommendation & 1Y target = Yahoo Finance analyst consensus, not financial advice."
            ).font = Font(name="Arial", italic=True, size=9)

    wb.save(out_path)
    return out_path


def main() -> int:
    tickers = load_universe()
    rows = scan(tickers)
    if not rows:
        print(f"No stocks with RSI >= {RSI_THRESHOLD} today. No report generated.")
        return 0
    path = write_excel(rows)
    print(f"\n{len(rows)} stocks flagged. Report saved to: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())


# ---------------------------------------------------------------------------
# SCHEDULING THE AGENT (daily, ~09:30 IST after NSE opens at 09:15)
#
# macOS / Linux (cron):
#   crontab -e
#   30 9 * * 1-5  /usr/bin/python3 /path/to/nifty_rsi_screener.py >> ~/rsi_reports/agent.log 2>&1
#
# Windows (Task Scheduler):
#   Create Basic Task -> Daily 9:30 AM -> Start a program:
#     Program: python    Arguments: C:\path\to\nifty_rsi_screener.py
#
# Tip: RSI is computed on daily closes, so a scan at 09:30 uses yesterday's
# completed candle plus today's live price. For a pure end-of-day view,
# schedule it after 15:30 IST instead.
# ---------------------------------------------------------------------------
