"""
Momentum Desk - NSE RSI Screener
--------------------------------
Daily screen of NSE stocks by 14-day RSI, with a per-stock detail view
covering market cap, support/resistance and a financial health scorecard.

Deploy free on Streamlit Community Cloud (share.streamlit.io). Locally:
  pip install -r requirements.txt
  streamlit run streamlit_app.py
"""

from __future__ import annotations

import datetime as dt
import io
import time
from collections import Counter

import pandas as pd
import requests
import streamlit as st
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RSI_PERIOD = 14
MCAP_LARGE_CR, MCAP_MID_CR, MCAP_SMALL_CR = 20000, 5000, 500

NIFTY50 = [
    "ADANIENT.NS", "ADANIPORTS.NS", "APOLLOHOSP.NS", "ASIANPAINT.NS",
    "AXISBANK.NS", "BAJAJ-AUTO.NS", "BAJFINANCE.NS", "BAJAJFINSV.NS",
    "BEL.NS", "BHARTIARTL.NS", "CIPLA.NS", "COALINDIA.NS", "DRREDDY.NS",
    "EICHERMOT.NS", "GRASIM.NS", "HCLTECH.NS", "HDFCBANK.NS",
    "HDFCLIFE.NS", "HEROMOTOCO.NS", "HINDALCO.NS", "HINDUNILVR.NS",
    "ICICIBANK.NS", "INDUSINDBK.NS", "INFY.NS", "ITC.NS", "JSWSTEEL.NS",
    "KOTAKBANK.NS", "LT.NS", "M&M.NS", "MARUTI.NS", "NESTLEIND.NS",
    "NTPC.NS", "ONGC.NS", "POWERGRID.NS", "RELIANCE.NS", "SBILIFE.NS",
    "SBIN.NS", "SHRIRAMFIN.NS", "SUNPHARMA.NS", "TATACONSUM.NS",
    "TATAMOTORS.NS", "TATASTEEL.NS", "TCS.NS", "TECHM.NS", "TITAN.NS",
    "TRENT.NS", "ULTRACEMCO.NS", "WIPRO.NS",
]

HEADERS = [
    "Date", "Stock Symbol", "Sector", "Current Price (Rs)", "52 Week High (Rs)",
    "RSI", "Buy/Sell", "1 Year Target (Rs)",
]

_EQUITY_LIST_URL = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
_INDEX500_LIST_URL = "https://niftyindices.com/IndexConstituent/ind_nifty500list.csv"

DISCLAIMER = (
    "Everything shown here is for reference and general information only. It is not "
    "investment advice, not a recommendation to buy or sell any security, and not a "
    "solicitation of any kind. The Buy/Sell rating and 1-year target are third-party "
    "analyst consensus figures reproduced as-is, not the view of this app. Technical "
    "levels and health checks are mechanically computed from reported financials and "
    "may be wrong, incomplete or stale. Do your own research and consult a "
    "SEBI-registered adviser before investing. Securities investments carry market "
    "risk, including possible loss of capital."
)

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Archivo:wght@500;600;700&family=Inter:wght@400;500&family=IBM+Plex+Mono:wght@400;500;600&display=swap');

:root{
  --ink:#101A24; --panel:#F4F7F9; --line:#DFE7ED;
  --primary:#0E7C86; --primary-soft:#E3F2F3;
  --pos:#0B7A4B; --neg:#B3261E; --neu:#6B7A8F; --nodata:#A8B4BF;
  --text:#16202B; --muted:#5E6E7E;
}
html, body, [class*="css"], .stMarkdown { font-family:'Inter', sans-serif; color:var(--text); }
h1,h2,h3,h4,h5 { font-family:'Archivo', sans-serif; letter-spacing:-0.02em; }
[data-testid="stMetricValue"], [data-testid="stDataFrame"] { font-family:'IBM Plex Mono', monospace; }

.band{
  background:var(--ink); border-radius:14px; padding:22px 26px; margin-bottom:18px;
  display:flex; align-items:flex-end; justify-content:space-between; flex-wrap:wrap; gap:12px;
}
.band-name{ font-family:'Archivo',sans-serif; font-weight:700; font-size:1.55rem;
  color:#FFF; letter-spacing:-0.03em; line-height:1.1; }
.band-sub{ font-family:'IBM Plex Mono',monospace; font-size:0.78rem; color:#8FA3B4;
  margin-top:6px; letter-spacing:0.02em; }
.band-date{ font-family:'IBM Plex Mono',monospace; font-size:0.78rem; color:#C9D6E0;
  border:1px solid #2C3B49; border-radius:999px; padding:5px 12px; }

.stat{ background:var(--panel); border:1px solid var(--line); border-radius:12px;
  padding:14px 16px; height:100%; }
.stat-k{ font-size:0.7rem; text-transform:uppercase; letter-spacing:0.08em;
  color:var(--muted); font-weight:500; }
.stat-v{ font-family:'IBM Plex Mono',monospace; font-size:1.32rem; font-weight:600;
  color:var(--text); margin-top:4px; line-height:1.2; }
.stat-n{ font-size:0.75rem; color:var(--muted); margin-top:2px; }

.score-row{ display:flex; gap:10px; flex-wrap:wrap; margin:4px 0 14px; }
.score{ flex:1 1 120px; background:#FFF; border:1px solid var(--line);
  border-left:4px solid var(--neu); border-radius:10px; padding:12px 14px; }
.score.pos{ border-left-color:var(--pos); }
.score.neg{ border-left-color:var(--neg); }
.score.neu{ border-left-color:var(--neu); }
.score.nod{ border-left-color:var(--nodata); }
.score-n{ font-family:'IBM Plex Mono',monospace; font-size:1.6rem; font-weight:600; line-height:1; }
.score-l{ font-size:0.74rem; color:var(--muted); margin-top:5px; }

.chk{ display:flex; align-items:baseline; gap:10px; padding:7px 0;
  border-bottom:1px solid var(--line); font-size:0.87rem; }
.chk:last-child{ border-bottom:none; }
.tag{ font-family:'IBM Plex Mono',monospace; font-size:0.66rem; font-weight:600;
  padding:2px 7px; border-radius:4px; text-transform:uppercase; letter-spacing:0.05em;
  flex:0 0 auto; min-width:62px; text-align:center; }
.tag.Yes{ background:#E4F3EC; color:var(--pos); }
.tag.No{ background:#FBE7E5; color:var(--neg); }
.tag.Neutral{ background:#ECF0F3; color:var(--neu); }
.tag.NoData{ background:#F2F5F7; color:var(--nodata); }
.chk-n{ flex:1 1 auto; }
.chk-d{ font-family:'IBM Plex Mono',monospace; font-size:0.78rem; color:var(--muted);
  flex:0 0 auto; }

.lvl{ display:flex; justify-content:space-between; padding:6px 10px; border-radius:6px;
  font-family:'IBM Plex Mono',monospace; font-size:0.85rem; margin-bottom:3px; }
.lvl.r{ background:#FBECEA; color:#8C2F26; }
.lvl.p{ background:var(--primary-soft); color:#0A5A62; font-weight:600; }
.lvl.s{ background:#E7F3EC; color:#0A5C3A; }
.lvl.now{ background:var(--ink); color:#FFF; font-weight:600; }

.disc{ font-size:0.76rem; color:var(--muted); line-height:1.6;
  border-top:1px solid var(--line); padding-top:14px; margin-top:8px; }
.stButton>button{ border-radius:8px; font-weight:500; }
</style>
"""


# ------------------------------ indicators --------------------------------

def compute_rsi(close: pd.Series, period: int = RSI_PERIOD) -> float | None:
    close = close.dropna()
    if len(close) < period + 1:
        return None
    delta = close.diff()
    gain = delta.clip(lower=0.0)
    loss = -delta.clip(upper=0.0)
    avg_gain = gain.ewm(alpha=1.0 / period, min_periods=period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1.0 / period, min_periods=period, adjust=False).mean()
    g, l = avg_gain.iloc[-1], avg_loss.iloc[-1]
    if pd.isna(g) or pd.isna(l):
        return None
    if l == 0:
        return 100.0
    return float(100.0 - (100.0 / (1.0 + (g / l))))


def classify_market_cap(mcap_inr: float | None) -> str:
    if not mcap_inr or mcap_inr <= 0:
        return "n/a"
    cr = mcap_inr / 1e7
    if cr >= MCAP_LARGE_CR:
        return "Large cap"
    if cr >= MCAP_MID_CR:
        return "Mid cap"
    if cr >= MCAP_SMALL_CR:
        return "Small cap"
    return "Micro cap"


def pivot_levels(h: float, l: float, c: float) -> dict:
    p = (h + l + c) / 3.0
    return {"R3": h + 2 * (p - l), "R2": p + (h - l), "R1": 2 * p - l, "Pivot": p,
            "S1": 2 * p - h, "S2": p - (h - l), "S3": l - 2 * (h - p)}


def swing_levels(close: pd.Series, price: float, window: int = 10):
    s = close.dropna()
    highs = s[(s.shift(window) < s) & (s.shift(-window) < s)]
    lows = s[(s.shift(window) > s) & (s.shift(-window) > s)]
    sup, res = lows[lows < price], highs[highs > price]
    return (float(sup.max()) if len(sup) else None,
            float(res.min()) if len(res) else None)


# --------------------------- health scorecard -----------------------------

def _find_row(fin, *keys):
    if fin is None or getattr(fin, "empty", True):
        return None
    for k in keys:
        for idx in fin.index:
            if k.lower() in str(idx).lower():
                return fin.loc[idx]
    return None


def _vals(series, n=1):
    if series is None:
        return None
    s = pd.Series(series).dropna()
    if s.empty:
        return None
    return s.sort_index(ascending=False).head(n).astype(float)


def _band(v, good, ok, higher_better=True):
    if v is None:
        return "No Data"
    if higher_better:
        return "Yes" if v >= good else ("Neutral" if v >= ok else "No")
    return "Yes" if v <= good else ("Neutral" if v <= ok else "No")


def health_checks(income, balance, cash) -> list[dict]:
    """15 mechanical checks on reported financials. Each -> Yes/Neutral/No/No Data."""
    out = []

    def add(name, verdict, detail=""):
        out.append({"name": name, "verdict": verdict, "detail": detail})

    ni = _find_row(income, "Net Income")
    cfo = _find_row(cash, "Operating Cash Flow", "Total Cash From Operating")
    n3, c3 = _vals(ni, 3), _vals(cfo, 3)
    if n3 is not None and c3 is not None and n3.sum() != 0:
        r = c3.sum() / n3.sum()
        add("Cash flow backs reported profit", _band(r, 0.8, 0.5), f"{r:.2f}x")
    else:
        add("Cash flow backs reported profit", "No Data")

    c1 = _vals(cfo, 1)
    add("Operating cash flow positive",
        "Yes" if c1 is not None and c1.iloc[0] > 0 else ("No" if c1 is not None else "No Data"),
        f"Rs {c1.iloc[0]/1e7:,.0f} Cr" if c1 is not None else "")

    cx = _vals(_find_row(cash, "Capital Expenditure"), 1)
    if c1 is not None and cx is not None:
        fcf = c1.iloc[0] - abs(cx.iloc[0])
        add("Free cash flow positive", "Yes" if fcf > 0 else "No", f"Rs {fcf/1e7:,.0f} Cr")
    else:
        add("Free cash flow positive", "No Data")

    td = _vals(_find_row(balance, "Total Debt"), 2)
    eq = _vals(_find_row(balance, "Stockholders Equity", "Total Equity"), 1)
    if td is not None and eq is not None and eq.iloc[0]:
        de = td.iloc[0] / eq.iloc[0]
        add("Debt to equity", _band(de, 0.5, 1.5, False), f"{de:.2f}x")
    else:
        add("Debt to equity", "No Data")

    if td is not None and len(td) == 2:
        add("Debt reduced vs last year", "Yes" if td.iloc[0] < td.iloc[1] else "No",
            f"{td.iloc[0]/1e7:,.0f} Cr")
    else:
        add("Debt reduced vs last year", "No Data")

    ebit = _vals(_find_row(income, "EBIT", "Operating Income"), 1)
    ie = _vals(_find_row(income, "Interest Expense"), 1)
    if ebit is not None and ie is not None and ie.iloc[0]:
        cov = ebit.iloc[0] / abs(ie.iloc[0])
        add("Interest coverage", _band(cov, 5, 2), f"{cov:.1f}x")
    else:
        add("Interest coverage", "No Data")

    ca = _vals(_find_row(balance, "Current Assets"), 1)
    cl = _vals(_find_row(balance, "Current Liabilities"), 1)
    if ca is not None and cl is not None and cl.iloc[0]:
        cr = ca.iloc[0] / cl.iloc[0]
        add("Current ratio", _band(cr, 1.5, 1.0), f"{cr:.2f}x")
    else:
        add("Current ratio", "No Data")

    rev = _vals(_find_row(income, "Total Revenue"), 3)
    rg = (rev.iloc[0] / rev.iloc[1] - 1) if rev is not None and len(rev) >= 2 and rev.iloc[1] else None
    for label, ser in (("Receivables in line with sales",
                        _vals(_find_row(balance, "Receivables", "Accounts Receivable"), 2)),
                       ("Inventory in line with sales",
                        _vals(_find_row(balance, "Inventory"), 2))):
        if ser is not None and len(ser) == 2 and ser.iloc[1] and rg is not None:
            g = ser.iloc[0] / ser.iloc[1] - 1
            add(label, _band(g - rg, 0.10, 0.25, False), f"{g*100:.0f}% vs {rg*100:.0f}%")
        else:
            add(label, "No Data")

    if rev is not None and len(rev) >= 3 and rev.iloc[2]:
        add("Revenue growing (3y)", "Yes" if rev.iloc[0] > rev.iloc[2] else "No",
            f"{(rev.iloc[0]/rev.iloc[2]-1)*100:.0f}%")
    else:
        add("Revenue growing (3y)", "No Data")

    if n3 is not None and len(n3) >= 3 and n3.iloc[2]:
        add("Profit growing (3y)", "Yes" if n3.iloc[0] > n3.iloc[2] else "No",
            f"{(n3.iloc[0]/n3.iloc[2]-1)*100:.0f}%")
        add("Profitable every year (3y)", "Yes" if (n3 > 0).all() else "No")
    else:
        add("Profit growing (3y)", "No Data")
        add("Profitable every year (3y)", "No Data")

    if n3 is not None and eq is not None and eq.iloc[0]:
        roe = n3.iloc[0] / eq.iloc[0] * 100
        add("Return on equity", _band(roe, 15, 8), f"{roe:.1f}%")
    else:
        add("Return on equity", "No Data")

    om = _vals(_find_row(income, "Operating Income", "EBIT"), 2)
    if (om is not None and rev is not None and len(om) == 2 and len(rev) >= 2
            and rev.iloc[0] and rev.iloc[1]):
        m0, m1 = om.iloc[0] / rev.iloc[0], om.iloc[1] / rev.iloc[1]
        add("Operating margin holding up", "Yes" if m0 >= m1 else "No",
            f"{m0*100:.1f}% vs {m1*100:.1f}%")
    else:
        add("Operating margin holding up", "No Data")

    if n3 is not None and rev is not None and rev.iloc[0]:
        nm = n3.iloc[0] / rev.iloc[0] * 100
        add("Net margin positive", "Yes" if nm > 0 else "No", f"{nm:.1f}%")
    else:
        add("Net margin positive", "No Data")

    return out


# ------------------------------- universe ---------------------------------

def _fetch_symbol_csv(url: str) -> list[str]:
    hdr = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Accept": "text/csv,*/*"}
    try:
        r = requests.get(url, headers=hdr, timeout=20)
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text))
        col = next(c for c in df.columns if c.strip().upper() == "SYMBOL")
        return [f"{s.strip()}.NS" for s in df[col].dropna().astype(str) if s.strip()]
    except Exception:
        return []


@st.cache_data(ttl=86400, show_spinner=False)
def load_all_nse_tickers() -> list[str]:
    return _fetch_symbol_csv(_EQUITY_LIST_URL)


@st.cache_data(ttl=86400, show_spinner=False)
def load_nifty500_tickers() -> list[str]:
    return _fetch_symbol_csv(_INDEX500_LIST_URL)


def parse_uploaded_symbols(text: str) -> list[str]:
    out = []
    for line in text.replace(",", "\n").splitlines():
        s = line.strip().upper()
        if not s or s == "SYMBOL":
            continue
        out.append(s if s.endswith(".NS") else f"{s}.NS")
    return out


# ----------------------------- data fetching ------------------------------

def fetch_fundamentals(symbol: str) -> dict:
    name = symbol.replace(".NS", "")
    sector, reco, target = "n/a", "n/a", "n/a"
    tk = yf.Ticker(symbol)
    try:
        info = tk.info or {}
        name = info.get("shortName") or info.get("longName") or name
        if info.get("sector"):
            sector = str(info["sector"]).strip()
        rk = info.get("recommendationKey")
        if rk and rk.lower() != "none":
            reco = rk.replace("_", " ").title()
        tp = info.get("targetMeanPrice")
        if tp:
            target = round(float(tp), 2)
    except Exception:
        pass
    if target == "n/a":
        try:
            m = (tk.analyst_price_targets or {}).get("mean")
            if m:
                target = round(float(m), 2)
        except Exception:
            pass
    return {"name": name, "sector": sector, "reco": reco, "target": target}


@st.cache_data(ttl=1800, show_spinner=False)
def fetch_detail(symbol: str) -> dict:
    out = {"name": symbol.replace(".NS", ""), "sector": "n/a", "mcap": None,
           "price": None, "high52": None, "low52": None, "rsi": None,
           "reco": "n/a", "target": "n/a", "levels": {}, "swing": (None, None),
           "checks": [], "error": None}
    try:
        tk = yf.Ticker(symbol)
        hist = tk.history(period="2y", interval="1d", auto_adjust=False)
        if hist is None or hist.empty:
            out["error"] = "No price history found for this symbol. Check the spelling."
            return out
        close = hist["Close"].dropna()
        out["price"] = float(close.iloc[-1])
        out["rsi"] = compute_rsi(close)
        yr = hist.tail(252)
        out["high52"], out["low52"] = float(yr["High"].max()), float(yr["Low"].min())

        monthly = hist.resample("ME").agg({"High": "max", "Low": "min", "Close": "last"}).dropna()
        if len(monthly) >= 2:
            prev = monthly.iloc[-2]
            out["levels"] = pivot_levels(float(prev["High"]), float(prev["Low"]),
                                         float(prev["Close"]))
        out["swing"] = swing_levels(close.tail(180), out["price"])

        try:
            info = tk.info or {}
            out["name"] = info.get("shortName") or info.get("longName") or out["name"]
            out["mcap"] = info.get("marketCap")
            if info.get("sector"):
                out["sector"] = str(info["sector"]).strip()
            rk = info.get("recommendationKey")
            if rk and rk.lower() != "none":
                out["reco"] = rk.replace("_", " ").title()
            tp = info.get("targetMeanPrice")
            if tp:
                out["target"] = round(float(tp), 2)
        except Exception:
            pass
        try:
            out["checks"] = health_checks(tk.income_stmt, tk.balance_sheet, tk.cashflow)
        except Exception:
            out["checks"] = []
    except Exception as exc:
        out["error"] = f"Could not load this stock ({type(exc).__name__}). Try again shortly."
    return out


@st.cache_data(ttl=1800, show_spinner=False)
def scan(tickers: tuple[str, ...], threshold: float, fetch_fund_limit: int) -> pd.DataFrame:
    bar = st.progress(0.0, text="Downloading price history...")
    close_map, high_map = {}, {}
    batch, n = 200, len(tickers)
    for i in range(0, n, batch):
        chunk = list(tickers[i:i + batch])
        data = yf.download(chunk, period="1y", interval="1d", group_by="ticker",
                           auto_adjust=False, threads=True, progress=False)
        for sym in chunk:
            try:
                df = data[sym] if len(chunk) > 1 else data
                c = df["Close"].dropna()
                if c.empty:
                    continue
                close_map[sym] = c
                high_map[sym] = float(df["High"].dropna().max())
            except Exception:
                continue
        bar.progress(min((i + batch) / n, 1.0), text=f"Downloaded {min(i+batch, n)}/{n} stocks")

    hits = []
    for sym, c in close_map.items():
        rsi = compute_rsi(c)
        if rsi is not None and rsi >= threshold:
            hits.append((sym, rsi, float(c.iloc[-1]), high_map.get(sym)))
    hits.sort(key=lambda x: x[1], reverse=True)

    rows, total = [], min(len(hits), fetch_fund_limit)
    bar.progress(0.0, text=f"Loading details for {total} matches")
    for idx, (sym, rsi, price, high52) in enumerate(hits):
        if idx < fetch_fund_limit:
            f = fetch_fundamentals(sym)
            time.sleep(0.4)
            bar.progress((idx + 1) / max(total, 1), text=f"Loaded {idx+1}/{total}")
        else:
            f = {"name": sym.replace(".NS", ""), "sector": "not loaded",
                 "reco": "not loaded", "target": "not loaded"}
        rows.append({
            "Date": dt.date.today().isoformat(),
            "Stock Symbol": sym.replace(".NS", ""),
            "Sector": f["sector"],
            "Current Price (Rs)": round(price, 2),
            "52 Week High (Rs)": round(high52, 2) if high52 else None,
            "RSI": round(rsi, 1),
            "Buy/Sell": f["reco"],
            "1 Year Target (Rs)": f["target"],
        })
    bar.empty()
    return pd.DataFrame(rows, columns=HEADERS)


# --------------------------------- excel ----------------------------------

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RSI Screener"
    hfill = PatternFill("solid", fgColor="101A24")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    bfont = Font(name="Arial", size=11)
    hot = PatternFill("solid", fgColor="E3F2F3")

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for _, r in df.iterrows():
        ws.append([r[h] for h in HEADERS])
        i = ws.max_row
        for cell in ws[i]:
            cell.font = bfont
        for col in (4, 5, 8):
            c = ws.cell(row=i, column=col)
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0.00"
        ws.cell(row=i, column=6).number_format = "0.0"
        if isinstance(r["RSI"], (int, float)) and r["RSI"] >= 70:
            for cell in ws[i]:
                cell.fill = hot

    for i, w in enumerate([12, 16, 22, 16, 16, 8, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(ws.max_row, 2)}"
    ws.cell(row=ws.max_row + 2, column=1, value="Disclaimer: " + DISCLAIMER).font = \
        Font(name="Arial", italic=True, size=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------ detail view -------------------------------

def _stat(label: str, value: str, note: str = "") -> str:
    note_html = f'<div class="stat-n">{note}</div>' if note else ""
    return (f'<div class="stat"><div class="stat-k">{label}</div>'
            f'<div class="stat-v">{value}</div>{note_html}</div>')


def render_detail(symbol: str):
    d = fetch_detail(f"{symbol}.NS")
    if d["error"]:
        st.warning(d["error"])
        return

    st.markdown(f"### {symbol}")
    st.caption(f"{d['name']} · {d['sector']}")

    price = f"Rs {d['price']:,.2f}" if d["price"] else "n/a"
    mcap = f"Rs {d['mcap']/1e7:,.0f} Cr" if d["mcap"] else "n/a"
    rsi = f"{d['rsi']:.1f}" if d["rsi"] else "n/a"
    rng = f"{d['low52']:,.0f} – {d['high52']:,.0f}" if d["high52"] else "n/a"
    tgt = f"Rs {d['target']:,.2f}" if isinstance(d["target"], (int, float)) else "n/a"

    cols = st.columns(5)
    cards = [
        _stat("Price", price),
        _stat("Market cap", mcap, classify_market_cap(d["mcap"])),
        _stat("RSI (14d)", rsi, "overbought zone" if d["rsi"] and d["rsi"] >= 70 else ""),
        _stat("52-week range", rng),
        _stat("1-year target", tgt, f"consensus: {d['reco']}"),
    ]
    for col, card in zip(cols, cards):
        col.markdown(card, unsafe_allow_html=True)

    st.write("")
    tab_levels, tab_health = st.tabs(["Price levels", "Financial health"])

    with tab_levels:
        left, right = st.columns([1, 1])
        with left:
            st.markdown("**Monthly pivot levels**")
            if d["levels"]:
                lv, p = d["levels"], d["price"]
                html, placed = [], False
                for k in ["R3", "R2", "R1", "Pivot", "S1", "S2", "S3"]:
                    if not placed and p and lv[k] < p:
                        html.append(f'<div class="lvl now"><span>Current price</span>'
                                    f'<span>{p:,.2f}</span></div>')
                        placed = True
                    cls = "p" if k == "Pivot" else ("r" if k.startswith("R") else "s")
                    html.append(f'<div class="lvl {cls}"><span>{k}</span>'
                                f'<span>{lv[k]:,.2f}</span></div>')
                if not placed and p:
                    html.append(f'<div class="lvl now"><span>Current price</span>'
                                f'<span>{p:,.2f}</span></div>')
                st.markdown("".join(html), unsafe_allow_html=True)
                st.caption("Floor-trader pivots from the last completed month.")
            else:
                st.write("Not enough history to compute pivots.")
        with right:
            st.markdown("**Nearest swing levels**")
            sup, res = d["swing"]
            st.markdown(
                f'<div class="lvl r"><span>Resistance</span><span>'
                f'{f"{res:,.2f}" if res else "none in range"}</span></div>'
                f'<div class="lvl now"><span>Current price</span><span>'
                f'{d["price"]:,.2f}</span></div>'
                f'<div class="lvl s"><span>Support</span><span>'
                f'{f"{sup:,.2f}" if sup else "none in range"}</span></div>',
                unsafe_allow_html=True)
            st.caption("Local highs and lows over roughly the last six months.")

    with tab_health:
        checks = d["checks"]
        if not checks:
            st.write("No financial statements available for this stock.")
        else:
            cnt = Counter(c["verdict"] for c in checks)
            st.markdown(
                '<div class="score-row">'
                f'<div class="score pos"><div class="score-n">{cnt.get("Yes",0)}</div>'
                '<div class="score-l">Positive</div></div>'
                f'<div class="score neu"><div class="score-n">{cnt.get("Neutral",0)}</div>'
                '<div class="score-l">Neutral</div></div>'
                f'<div class="score neg"><div class="score-n">{cnt.get("No",0)}</div>'
                '<div class="score-l">Negative</div></div>'
                f'<div class="score nod"><div class="score-n">{cnt.get("No Data",0)}</div>'
                '<div class="score-l">No data</div></div>'
                '</div>', unsafe_allow_html=True)
            rows = "".join(
                f'<div class="chk"><span class="tag {c["verdict"].replace(" ","")}">'
                f'{c["verdict"]}</span><span class="chk-n">{c["name"]}</span>'
                f'<span class="chk-d">{c["detail"]}</span></div>' for c in checks)
            st.markdown(rows, unsafe_allow_html=True)
            st.caption("Mechanical checks on reported annual financials. A negative flag "
                       "marks something worth investigating, not a verdict on the company. "
                       "Ratios like debt-to-equity do not carry the same meaning for banks "
                       "and NBFCs. Not covered: promoter pledging, auditor changes, "
                       "related-party transactions, contingent liabilities.")


@st.dialog("Stock detail", width="large")
def detail_dialog(symbol: str):
    render_detail(symbol)


# ---------------------------------- app -----------------------------------

st.set_page_config(page_title="Momentum Desk — NSE RSI Screener",
                   page_icon="📊", layout="wide")
st.markdown(CSS, unsafe_allow_html=True)

st.markdown(
    '<div class="band"><div><div class="band-name">Momentum Desk</div>'
    '<div class="band-sub">NSE · 14-day RSI screen · click any row for detail</div></div>'
    f'<div class="band-date">{dt.date.today().strftime("%d %b %Y")}</div></div>',
    unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### Screen")
    threshold = st.slider("RSI at or above", 50, 90, 65, 1)
    universe_choice = st.radio(
        "Stocks to scan",
        ["NIFTY 50 — fastest", "NIFTY 500 — recommended", "All NSE — slow", "My own list"])
    uploaded = None
    if universe_choice == "My own list":
        uploaded = st.text_area("One symbol per line", placeholder="RELIANCE\nTCS\nINFY")
    fetch_limit = st.number_input("Load details for top N matches", 10, 500, 100, 10,
                                  help="Each stock is a separate lookup. A lower number "
                                       "keeps big scans from timing out.")
    c_run, c_clear = st.columns(2)
    run = c_run.button("Run screen", type="primary", use_container_width=True)
    clear = c_clear.button("Clear", use_container_width=True)

    st.divider()
    st.markdown("### Look up a stock")
    all_syms = load_all_nse_tickers()
    if all_syms:
        opts = ["—"] + [s.replace(".NS", "") for s in all_syms]
        picked_search = st.selectbox("Search by symbol", opts,
                                     help="Type to filter the full NSE list.")
        if picked_search != "—":
            if st.button("Open detail", use_container_width=True):
                detail_dialog(picked_search)
    else:
        typed = st.text_input("Symbol", placeholder="RELIANCE")
        if st.button("Open detail", use_container_width=True) and typed.strip():
            detail_dialog(typed.strip().upper())

if clear:
    for k in ("results", "scanned", "threshold", "opened_for", "screener_table"):
        st.session_state.pop(k, None)
    st.rerun()

if run:
    if universe_choice == "NIFTY 50 — fastest":
        tickers = NIFTY50
    elif universe_choice == "NIFTY 500 — recommended":
        tickers = load_nifty500_tickers()
        if not tickers:
            st.error("The NIFTY 500 list did not load. Try again in a minute, or paste "
                     "your own list in the sidebar.")
            st.stop()
    elif universe_choice == "All NSE — slow":
        tickers = load_all_nse_tickers()
        if not tickers:
            st.error("The NSE list did not load. Try again in a minute, or paste your "
                     "own list in the sidebar.")
            st.stop()
    else:
        tickers = parse_uploaded_symbols(uploaded or "")
        if not tickers:
            st.warning("Add at least one symbol to scan.")
            st.stop()
    st.session_state["results"] = scan(tuple(tickers), float(threshold), int(fetch_limit))
    st.session_state["scanned"] = len(tickers)
    st.session_state["threshold"] = threshold
    st.session_state["opened_for"] = None

results = st.session_state.get("results")

if results is None:
    st.info("Set your RSI threshold and stock list in the sidebar, then run the screen. "
            "You can also look up any single stock from there.")
elif results.empty:
    st.warning(f"Nothing is at RSI {st.session_state.get('threshold', threshold)} or above "
               "right now. Lower the threshold or widen the stock list.")
else:
    st.markdown(f"**{len(results)} stocks** at RSI "
                f"{st.session_state.get('threshold', threshold)} or above, "
                f"from {st.session_state.get('scanned', 0)} scanned.")
    event = st.dataframe(
        results, use_container_width=True, hide_index=True,
        on_select="rerun", selection_mode="single-row", key="screener_table",
        column_config={
            "RSI": st.column_config.ProgressColumn("RSI", min_value=0, max_value=100,
                                                   format="%.1f"),
            "Current Price (Rs)": st.column_config.NumberColumn(format="%.2f"),
            "52 Week High (Rs)": st.column_config.NumberColumn(format="%.2f"),
        })

    picked = []
    try:
        picked = list(event.selection.rows)
    except Exception:
        pass
    if picked:
        sym = str(results.iloc[picked[0]]["Stock Symbol"])
        if st.session_state.get("opened_for") != sym:
            st.session_state["opened_for"] = sym
            detail_dialog(sym)
    else:
        st.session_state["opened_for"] = None

    st.download_button("Download Excel report", data=to_excel_bytes(results),
                       file_name=f"RSI_Screener_{dt.date.today().isoformat()}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")
    na = int((results["Buy/Sell"] == "n/a").sum())
    if na:
        st.caption(f"{na} of {len(results)} stocks have no published analyst rating — "
                   "common for smaller companies.")

st.markdown(f'<div class="disc"><strong>Disclaimer</strong> — {DISCLAIMER}</div>',
            unsafe_allow_html=True)
